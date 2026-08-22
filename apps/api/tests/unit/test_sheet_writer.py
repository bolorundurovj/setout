"""The workbook Setout writes, which its own reader has to recognise."""

from __future__ import annotations

import io
from datetime import date, datetime

import pytest
from openpyxl import load_workbook

from setout.services.sheets import read_budget, read_expenses, read_outstanding, write
from setout.services.sheets.detect import Found, Shape, find
from setout.services.sheets.workbook import read

pytestmark = pytest.mark.unit


def a_book() -> write.Book:
    return write.Book(
        project_name="Jacaranda Close, Ewuru",
        exponent=2,
        scopes=[
            write.PlanScope(
                code="1000",
                name="Administrative Expenses",
                lines=[
                    write.PlanLine("1001", "Beacons", 1_000_00, "fixed"),
                    write.PlanLine("1002", "Drawings", 20_000_00, "material"),
                ],
            )
        ],
        vendors=[
            write.VendorRow(
                name="Bright Star Aluminium",
                trade="Roofing",
                contact_name="Mr Sola",
                phone="0800 000 0001",
                email="a@b.com",
                notes="",
            )
        ],
        spend=[
            write.SpendRow(
                vendor="Bright Star Aluminium",
                code="1001",
                description="Roofing sheets",
                spent_on=date(2026, 8, 14),
                amount=500_000_00,
                document="receipt.png",
                paid_by="Ayo",
                notes="@ rate of 4200",
            )
        ],
        owed=[
            write.OwedRow(
                description="17 bags of cement",
                vendor="Corner Depot",
                promised="this week",
                received_on=None,
            )
        ],
        people=[write.PersonRow(name="Aunty Ngozi", role="Site keeper", phone="", notes="")],
        agreements=[
            write.AgreementRow(
                vendor="Kunle Bricklaying",
                description="Block work",
                agreed_amount=180_000_00,
                paid_amount=145_000_00,
            )
        ],
        prices=[
            write.PriceRow(
                name="Cement",
                unit="per bag",
                unit_rate=7_650_00,
                vendor="Corner Depot",
                spent_on=date(2026, 8, 14),
            )
        ],
        files=[
            write.FileRow(
                filename="receipt.png", against="Roofing sheets", byte_size=2048, checksum="ab" * 32
            )
        ],
    )


def sheets_of(data: bytes) -> dict[str, list[list[object]]]:
    book = load_workbook(io.BytesIO(data), data_only=True)
    return {
        name: [list(row) for row in book[name].iter_rows(values_only=True)]
        for name in book.sheetnames
    }


def test_the_reader_recognises_every_sheet_the_writer_names() -> None:
    found = find(read(write.build(a_book()), "book.xlsx"))

    assert {hit.shape for hit in found.found} == {
        Shape.BUDGET,
        Shape.VENDORS,
        Shape.EXPENSES,
        Shape.OUTSTANDING,
    }


def test_a_scope_heading_carries_the_total_of_its_lines() -> None:
    rows = sheets_of(write.build(a_book()))["Budget"]

    heading = next(row for row in rows if row[0] == "1000")
    lines = [row for row in rows if str(row[0]).startswith("100") and row[0] != "1000"]

    assert heading[5] == 21_000
    assert [line[5] for line in lines] == [1_000, 20_000]


def test_figures_are_written_as_the_money_they_are() -> None:
    rows = sheets_of(write.build(a_book()))["Invoices and Purchases"]
    row = rows[-1]

    assert row[5] == 500_000
    assert row[3] == "Roofing sheets"


def test_the_sheets_setout_alone_keeps_are_there_too() -> None:
    written = sheets_of(write.build(a_book()))

    assert "Aunty Ngozi" in [row[0] for row in written["People"]]
    assert "Kunle Bricklaying" in [row[0] for row in written["Agreements"]]
    assert "Cement" in [row[0] for row in written["Item prices"]]
    assert "receipt.png" in [row[0] for row in written["Attachments"]]


def test_an_export_of_nothing_is_still_a_workbook_the_reader_takes() -> None:
    empty = write.Book(project_name="Nothing yet", exponent=2)

    found = find(read(write.build(empty), "book.xlsx"))

    assert {hit.shape for hit in found.found} == {
        Shape.BUDGET,
        Shape.VENDORS,
        Shape.EXPENSES,
        Shape.OUTSTANDING,
    }


@pytest.mark.parametrize(
    ("name", "expected"),
    [
        ("Jacaranda Close, Ewuru", "setout-jacaranda-close-ewuru-20260819.xlsx"),
        ("  ", "setout-project-20260819.xlsx"),
        ("../etc/passwd", "setout-etcpasswd-20260819.xlsx"),
    ],
)
def test_the_filename_is_made_of_the_project_name(name: str, expected: str) -> None:
    assert write.filename_for(name, datetime(2026, 8, 19)) == expected


class TestTheSamplesMatchAnExport:
    def sheets_in(self, data: bytes) -> list[str]:
        return load_workbook(io.BytesIO(data)).sheetnames

    def test_a_sample_carries_the_sheets_an_export_does(self) -> None:
        export = self.sheets_in(write.build(a_book()))

        for filled in (False, True):
            sample = self.sheets_in(write.template(filled=filled))
            assert sample[0] == "How to fill this in"
            assert sample[1:] == export

    def test_the_filled_sample_shows_every_sheet_with_something_in_it(self) -> None:
        written = sheets_of(write.template(filled=True))

        for name in ["Budget", "Invoices and Purchases", "People", "Agreements", "Item prices"]:
            assert len(written[name]) > 3, name

    def test_the_blank_sample_carries_headings_and_no_rows(self) -> None:
        written = sheets_of(write.template())

        for name in ["Budget", "Invoices and Purchases", "Outstanding"]:
            assert len(written[name]) == 3, name


class TestWhatSetoutWroteReadsBack:
    def read_back(self, book: write.Book) -> dict[Shape, Found]:
        found = find(read(write.build(book), "book.xlsx"))
        return {hit.shape: hit for hit in found.found}

    def test_something_delivered_does_not_come_back_as_still_waiting(self) -> None:
        book = write.Book(
            project_name="P",
            exponent=2,
            owed=[
                write.OwedRow("20 rods", "A vendor", "this week", date(2026, 8, 14)),
                write.OwedRow("17 bags", "A vendor", "next week", None),
            ],
        )

        rows = read_outstanding.read(self.read_back(book)[Shape.OUTSTANDING]).owed

        assert [(row.description, row.resolved) for row in rows] == [
            ("20 rods", True),
            ("17 bags", False),
        ]

    def test_the_figures_come_back_as_the_figures_that_went_in(self) -> None:
        book = a_book()
        sheets = self.read_back(book)

        planned = read_budget.read(sheets[Shape.BUDGET], 2)
        spend = read_expenses.read(sheets[Shape.EXPENSES], 2)

        assert planned.planned_amount == book.scopes[0].planned_amount
        assert [line.cost_type for line in planned.scopes[0].lines] == ["fixed", "material"]
        assert [row.amount for row in spend.spend] == [book.spend[0].amount]
        assert spend.spend[0].description == "Roofing sheets"
