from __future__ import annotations

import csv
import io
from dataclasses import dataclass, field

from openpyxl import load_workbook

MAX_SHEETS = 24
MAX_ROWS = 20_000


@dataclass
class Sheet:
    name: str
    rows: list[list[object]] = field(default_factory=list)


class SheetError(Exception):
    """The file could not be opened as a spreadsheet."""


def read(data: bytes, filename: str) -> list[Sheet]:
    if filename.lower().endswith(".csv"):
        return [_csv(data, filename)]
    return _xlsx(data)


def _csv(data: bytes, filename: str) -> Sheet:
    try:
        text = data.decode("utf-8-sig")
    except UnicodeDecodeError:
        text = data.decode("latin-1")
    rows: list[list[object]] = [list(row) for row in csv.reader(io.StringIO(text))]
    stem = filename.rsplit("/", 1)[-1].rsplit(".", 1)[0]
    return Sheet(name=stem or "Sheet1", rows=rows[:MAX_ROWS])


def _xlsx(data: bytes) -> list[Sheet]:
    try:
        # data_only reads what the sheet last showed, not the formulas behind it.
        book = load_workbook(io.BytesIO(data), data_only=True, read_only=True)
    except Exception as exc:  # noqa: BLE001
        raise SheetError("That file could not be opened as a spreadsheet") from exc

    sheets: list[Sheet] = []
    for name in book.sheetnames[:MAX_SHEETS]:
        rows: list[list[object]] = []
        for row in book[name].iter_rows(values_only=True):
            rows.append(list(row))
            if len(rows) >= MAX_ROWS:
                break
        sheets.append(Sheet(name=name, rows=rows))
    book.close()
    return sheets


CURRENCY_LABELS = {"currency code", "currency"}


def currency_in(sheets: list[Sheet]) -> str:
    """The currency a Setout workbook names on its Project sheet, if it has one."""
    for sheet in sheets:
        for row in sheet.rows[:12]:
            for index, cell in enumerate(row[:-1]):
                label = str(cell or "").strip().lower().rstrip(":")
                if label in CURRENCY_LABELS:
                    code = str(row[index + 1] or "").strip().upper()
                    if len(code) == 3 and code.isalpha():
                        return code
    return ""
