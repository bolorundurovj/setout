"""Build the workbook a project came from, so nothing is trapped in Setout."""

from __future__ import annotations

import csv
import io
from dataclasses import dataclass, field
from datetime import date, datetime
from decimal import Decimal

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

TITLE_FONT = Font(name="Calibri", size=14, bold=True, color="1E2226")
HEAD_FONT = Font(name="Calibri", size=11, bold=True, color="1E2226")
HEAD_FILL = PatternFill("solid", fgColor="ECEFF3")
SCOPE_FONT = Font(name="Calibri", size=11, bold=True)
MONEY_FORMAT = "#,##0.00"
XLSX_TYPE = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
DATE_FORMAT = "dd mmm yyyy"

BUDGET_HEAD = [
    "Cost Codes",
    "Activities",
    "Labor Costs",
    "Material Costs",
    "Fixed Costs",
    "Budget",
]
VENDOR_HEAD = ["Company Name", "Trade", "Point of Contact", "Phone", "Email", "Notes"]
SPEND_HEAD = [
    "Invoice No.",
    "Subcontractors & Vendors",
    "Cost Code",
    "Description",
    "Date Invoiced",
    "Total Cost",
    "Documents",
    "Purchased By",
    "Paid",
    "Date Paid",
    "Notes",
]
OWED_HEAD = ["Item", "Vendor", "Qty", "Location", "status"]
PEOPLE_HEAD = ["Person", "Role", "Phone", "Notes"]
AGREEMENT_HEAD = ["Vendor", "Work agreed", "Agreed", "Paid", "Left", "Notes"]
PRICE_HEAD = ["Item name", "Unit", "Last price", "Bought from", "When"]
FILE_HEAD = ["File", "Kept against", "Size in bytes", "Checksum"]


COST_COLUMN = {"labour": 0, "material": 1, "fixed": 2}


@dataclass
class PlanLine:
    code: str
    name: str
    planned_amount: int
    cost_type: str | None = None


@dataclass
class PlanScope:
    code: str
    name: str
    lines: list[PlanLine] = field(default_factory=list)

    @property
    def planned_amount(self) -> int:
        return sum(line.planned_amount for line in self.lines)


@dataclass
class SpendRow:
    vendor: str
    code: str
    description: str
    spent_on: date | None
    amount: int
    document: str
    paid_by: str
    notes: str


@dataclass
class VendorRow:
    name: str
    trade: str
    contact_name: str
    phone: str
    email: str
    notes: str


@dataclass
class OwedRow:
    description: str
    vendor: str
    promised: str
    received_on: date | None


@dataclass
class PersonRow:
    name: str
    role: str
    phone: str
    notes: str


@dataclass
class AgreementRow:
    vendor: str
    description: str
    agreed_amount: int
    paid_amount: int


@dataclass
class PriceRow:
    name: str
    unit: str
    unit_rate: int | None
    vendor: str
    spent_on: date | None


@dataclass
class FileRow:
    filename: str
    against: str
    byte_size: int
    checksum: str


@dataclass
class Book:
    project_name: str
    exponent: int
    currency_code: str = ""
    scopes: list[PlanScope] = field(default_factory=list)
    vendors: list[VendorRow] = field(default_factory=list)
    spend: list[SpendRow] = field(default_factory=list)
    owed: list[OwedRow] = field(default_factory=list)
    people: list[PersonRow] = field(default_factory=list)
    agreements: list[AgreementRow] = field(default_factory=list)
    prices: list[PriceRow] = field(default_factory=list)
    files: list[FileRow] = field(default_factory=list)


def major(minor: int, exponent: int) -> Decimal:
    return Decimal(minor).scaleb(-exponent)


def compose(book: Book) -> Workbook:
    """One writer, so a sample cannot drift from an export."""
    wb = Workbook()
    wb.remove(wb.active)

    _project(wb.create_sheet("Project"), book)
    _budget(wb.create_sheet("Budget"), book)
    _vendors(wb.create_sheet("Vendors & Subcontractors"), book)
    _spend(wb.create_sheet("Invoices and Purchases"), book)
    _owed(wb.create_sheet("Outstanding"), book)
    _people(wb.create_sheet("People"), book)
    _agreements(wb.create_sheet("Agreements"), book)
    _prices(wb.create_sheet("Item prices"), book)
    _files(wb.create_sheet("Attachments"), book)
    return wb


def as_bytes(wb: Workbook) -> bytes:
    out = io.BytesIO()
    wb.save(out)
    return out.getvalue()


def build(book: Book) -> bytes:
    return as_bytes(compose(book))


CURRENCY_LABEL = "Currency code"


def _project(sheet: Worksheet, book: Book) -> None:
    sheet.append([f"{book.project_name}"])
    sheet.append([])
    for label, value in [
        ("Project", book.project_name),
        (CURRENCY_LABEL, book.currency_code),
        ("Amounts are written in", "whole units, not minor units"),
    ]:
        sheet.append([label, value])
    sheet["A1"].font = TITLE_FONT
    for row in sheet.iter_rows(min_row=3):
        row[0].font = HEAD_FONT
    _widths(sheet, [26, 44])


def _money_format(book: Book) -> str:
    return f'"{book.currency_code} "#,##0.00' if book.currency_code else MONEY_FORMAT


def _title(sheet: Worksheet, name: str, head: list[str]) -> None:
    sheet.append([name])
    sheet.append([])
    sheet.append(head)
    sheet["A1"].font = TITLE_FONT
    for column in range(1, len(head) + 1):
        cell = sheet.cell(row=3, column=column)
        cell.font = HEAD_FONT
        cell.fill = HEAD_FILL
        cell.alignment = Alignment(horizontal="left", vertical="center")
    sheet.freeze_panes = "A4"
    sheet.auto_filter.ref = f"A3:{get_column_letter(len(head))}3"


def _widths(sheet: Worksheet, widths: list[int]) -> None:
    for index, width in enumerate(widths, start=1):
        sheet.column_dimensions[get_column_letter(index)].width = width


def _format(
    sheet: Worksheet,
    money: list[int],
    dates: list[int] | None = None,
    money_format: str = MONEY_FORMAT,
) -> None:
    for row in sheet.iter_rows(min_row=4):
        for column in money:
            row[column].number_format = money_format
        for column in dates or []:
            row[column].number_format = DATE_FORMAT


def _budget(sheet: Worksheet, book: Book) -> None:
    _title(sheet, f"{book.project_name}: budget", BUDGET_HEAD)

    def figure(minor: int) -> Decimal:
        return major(minor, book.exponent)

    for scope in book.scopes:
        # Setout adds the lines up itself, so a re-import reads them alone.
        totals = _by_cost_type(scope.lines)
        sheet.append(
            [
                scope.code,
                scope.name,
                figure(totals["labour"]) if totals["labour"] else "",
                figure(totals["material"]) if totals["material"] else "",
                figure(totals["fixed"]) if totals["fixed"] else "",
                figure(scope.planned_amount),
            ]
        )
        sheet.cell(row=sheet.max_row, column=1).font = SCOPE_FONT
        sheet.cell(row=sheet.max_row, column=2).font = SCOPE_FONT
        sheet.cell(row=sheet.max_row, column=6).font = SCOPE_FONT
        for line in scope.lines:
            split: list[object] = ["", "", ""]
            if line.cost_type in COST_COLUMN:
                split[COST_COLUMN[line.cost_type]] = figure(line.planned_amount)
            sheet.append([line.code, line.name, *split, figure(line.planned_amount)])

    _widths(sheet, [14, 40, 15, 15, 15, 16])
    _format(sheet, money=[2, 3, 4, 5], money_format=_money_format(book))


def _by_cost_type(lines: list[PlanLine]) -> dict[str, int]:
    totals = {"labour": 0, "material": 0, "fixed": 0}
    for line in lines:
        if line.cost_type in totals:
            totals[line.cost_type] += line.planned_amount
    return totals


def _vendors(sheet: Worksheet, book: Book) -> None:
    _title(sheet, "Vendors & subcontractors", VENDOR_HEAD)
    for vendor in book.vendors:
        sheet.append(
            [
                vendor.name,
                vendor.trade,
                vendor.contact_name,
                vendor.phone,
                vendor.email,
                vendor.notes,
            ]
        )
    _widths(sheet, [28, 20, 22, 18, 26, 30])


def _spend(sheet: Worksheet, book: Book) -> None:
    _title(sheet, "Invoices & purchases", SPEND_HEAD)
    for row in book.spend:
        sheet.append(
            [
                "",
                row.vendor,
                row.code,
                row.description,
                row.spent_on,
                major(row.amount, book.exponent),
                row.document,
                row.paid_by,
                "Yes",
                row.spent_on,
                row.notes,
            ]
        )
    _widths(sheet, [12, 28, 12, 34, 14, 14, 18, 18, 8, 14, 30])
    _format(sheet, money=[5], dates=[4, 9], money_format=_money_format(book))


def _owed(sheet: Worksheet, book: Book) -> None:
    _title(sheet, "Paid for, not delivered", OWED_HEAD)
    for row in book.owed:
        arrived = f"Delivered {row.received_on:%d %b %Y}" if row.received_on else "Waiting"
        sheet.append([row.description, row.vendor, "", row.promised, arrived])
    _widths(sheet, [34, 26, 10, 22, 22])


def _people(sheet: Worksheet, book: Book) -> None:
    _title(sheet, "People", PEOPLE_HEAD)
    for person in book.people:
        sheet.append([person.name, person.role, person.phone, person.notes])
    _widths(sheet, [26, 20, 18, 30])


def _agreements(sheet: Worksheet, book: Book) -> None:
    _title(sheet, "Agreements", AGREEMENT_HEAD)
    for row in book.agreements:
        left = row.agreed_amount - row.paid_amount
        sheet.append(
            [
                row.vendor,
                row.description,
                major(row.agreed_amount, book.exponent),
                major(row.paid_amount, book.exponent),
                major(left, book.exponent),
                "",
            ]
        )
    _widths(sheet, [26, 34, 15, 15, 15, 30])
    _format(sheet, money=[2, 3, 4], money_format=_money_format(book))


def _prices(sheet: Worksheet, book: Book) -> None:
    _title(sheet, "Item prices", PRICE_HEAD)
    for row in book.prices:
        rate = "" if row.unit_rate is None else major(row.unit_rate, book.exponent)
        sheet.append([row.name, row.unit, rate, row.vendor, row.spent_on])
    _widths(sheet, [28, 16, 15, 26, 16])
    _format(sheet, money=[2], dates=[4], money_format=_money_format(book))


def _files(sheet: Worksheet, book: Book) -> None:
    _title(sheet, "Attachments, which are not in this workbook", FILE_HEAD)
    for row in book.files:
        sheet.append([row.filename, row.against, row.byte_size, row.checksum])
    _widths(sheet, [30, 34, 16, 66])


HOW_TO = [
    "A blank sheet to fill in and bring into Setout",
    "",
    "Each sheet below is recognised by its headings, so keep the heading row as it is.",
    "Add rows underneath. Delete a sheet you have nothing for.",
    "",
    "Budget: a cost code ending in 000 is a scope heading. Anything else is a line",
    "under the heading above it, so 3000 is a scope and 3001 sits inside it.",
    "Only the lines are read; a heading total is worked out from them.",
    "",
    "Put a planned figure in one of the labour, material or fixed columns and it is",
    "kept under that heading. The last column is the figure that counts either way.",
    "",
    "Invoices and Purchases: one row is one payment. The cost code says which scope",
    "it belongs to; leave it empty and the spending arrives unfiled, which is allowed.",
    "",
    "Amounts are plain numbers. No currency symbols, no thousands separators.",
    "Dates can be written any way a spreadsheet understands.",
    "",
    "Nothing here writes a budget from a spending row, or spending from a budget row.",
    "",
    "The last four sheets are what Setout adds when it exports a project. They are",
    "kept for you to read; nothing in them is brought back in, so leave them empty.",
]


def example_book(currency_code: str = "NGN") -> Book:
    return Book(
        project_name="Jacaranda Close, Ewuru",
        exponent=2,
        currency_code=currency_code,
        scopes=[
            PlanScope(
                code="1000",
                name="Administrative expenses",
                lines=[
                    PlanLine("1001", "Drawings and permits", 120_000_00, "fixed"),
                    PlanLine("1002", "Beacons", 30_000_00, "material"),
                ],
            ),
            PlanScope(
                code="2000",
                name="Concrete foundation",
                lines=[
                    PlanLine("2001", "Sand and cement", 450_000_00, "material"),
                    PlanLine("2002", "Block work", 180_000_00, "labour"),
                ],
            ),
        ],
        vendors=[
            VendorRow(
                name="Corner Depot Cement",
                trade="Building materials",
                contact_name="Mr Sola",
                phone="0800 000 0002",
                email="sales@cornerdepot.example",
                notes="Delivers within the day",
            ),
            VendorRow(
                name="Kunle Bricklaying",
                trade="Block work",
                contact_name="Kunle",
                phone="0800 000 0001",
                email="",
                notes="Agreed a price for the whole job",
            ),
        ],
        spend=[
            SpendRow(
                vendor="Corner Depot Cement",
                code="2001",
                description="17 bags of cement",
                spent_on=date(2026, 6, 11),
                amount=76_500_00,
                document="receipt-11jun.jpg",
                paid_by="Aunty Ngozi",
                notes="Bought on site",
            ),
            SpendRow(
                vendor="Kunle Bricklaying",
                code="2002",
                description="Block work, first payment",
                spent_on=date(2026, 7, 2),
                amount=60_000_00,
                document="",
                paid_by="",
                notes="",
            ),
            SpendRow(
                vendor="",
                code="",
                description="Nails and sundries",
                spent_on=date(2026, 7, 18),
                amount=8_400_00,
                document="",
                paid_by="",
                notes="No cost code, so this arrives unfiled",
            ),
        ],
        owed=[
            OwedRow(
                description="20 lengths of reinforcement",
                vendor="Corner Depot Cement",
                promised="this week",
                received_on=None,
            )
        ],
        people=[
            PersonRow(
                name="Aunty Ngozi",
                role="Buys on site",
                phone="0800 000 0003",
                notes="Holds money for small purchases",
            )
        ],
        agreements=[
            AgreementRow(
                vendor="Kunle Bricklaying",
                description="Block work, whole job",
                agreed_amount=180_000_00,
                paid_amount=60_000_00,
            )
        ],
        prices=[
            PriceRow(
                name="Cement",
                unit="per bag",
                unit_rate=4_500_00,
                vendor="Corner Depot Cement",
                spent_on=date(2026, 6, 11),
            )
        ],
        files=[
            FileRow(
                filename="receipt-11jun.jpg",
                against="17 bags of cement",
                byte_size=248_512,
                checksum="0" * 64,
            )
        ],
    )


def template(filled: bool = False) -> bytes:
    book = (
        example_book()
        if filled
        else Book(project_name="Your project", exponent=2, currency_code="put yours here")
    )
    wb = compose(book)

    guide = wb.create_sheet("How to fill this in", 0)
    for line in HOW_TO:
        guide.append([line])
    guide["A1"].font = TITLE_FONT
    _widths(guide, [96])

    return as_bytes(wb)


def as_csv(rows: list[list[object]]) -> bytes:
    out = io.StringIO()
    writer = csv.writer(out, lineterminator="\n")
    for row in rows:
        writer.writerow(["" if cell is None else cell for cell in row])
    return out.getvalue().encode("utf-8-sig")


def budget_csv(filled: bool = False) -> bytes:
    rows: list[list[object]] = [list(BUDGET_HEAD)]
    if filled:
        for scope in example_book().scopes:
            rows.append([scope.code, scope.name, "", "", "", major(scope.planned_amount, 2)])
            for line in scope.lines:
                split: list[object] = ["", "", ""]
                if line.cost_type in COST_COLUMN:
                    split[COST_COLUMN[line.cost_type]] = major(line.planned_amount, 2)
                rows.append([line.code, line.name, *split, major(line.planned_amount, 2)])
    return as_csv(rows)


def spend_csv(filled: bool = False) -> bytes:
    rows: list[list[object]] = [list(SPEND_HEAD)]
    if filled:
        for row in example_book().spend:
            rows.append(
                [
                    "",
                    row.vendor,
                    row.code,
                    row.description,
                    row.spent_on,
                    major(row.amount, 2),
                    row.document,
                    row.paid_by,
                    "Yes",
                    row.spent_on,
                    row.notes,
                ]
            )
    return as_csv(rows)


def filename_for(project_name: str, when: datetime) -> str:
    keep = [c if c.isalnum() or c in " -_" else "" for c in project_name]
    stem = "".join(keep).strip().replace(" ", "-").lower() or "project"
    return f"setout-{stem}-{when:%Y%m%d}.xlsx"


SAMPLES: dict[str, tuple[str, str]] = {
    "blank": ("setout-blank-sheet.xlsx", XLSX_TYPE),
    "example": ("setout-example-sheet.xlsx", XLSX_TYPE),
    "budget-csv": ("setout-budget-example.csv", "text/csv"),
    "spending-csv": ("setout-spending-example.csv", "text/csv"),
}


def sample(kind: str) -> bytes:
    if kind == "example":
        return template(filled=True)
    if kind == "budget-csv":
        return budget_csv(filled=True)
    if kind == "spending-csv":
        return spend_csv(filled=True)
    return template()
